import streamlit as st
import plotly_express as px
import pandas as pd
from streamlit_option_menu import option_menu
from openpyxl import load_workbook

# page config
st.set_page_config(
    page_title='Data visualizer',
    page_icon='📄',
    layout='wide'
)

#& Title
#st.title('Data Visualization App')

#! Add sidebar
st.sidebar.subheader('Data Visualizer')





#$ Load Data
#@st.cache_data
#def load_dataset():
#    if uploaded_file is not None:
#        data=pd.read_table(uploaded_file)
#        return data
#
#
#with st.spinner('Loading Dataset...'):
#    uploaded_file =load_dataset()
#    st.sidebar.success('Loaded Dataset')


with st.sidebar:
    selected = option_menu("Main Menu", ["Home",'File Upload', 'Settings'], 
        icons=['house','cloud-upload' ,'gear'], menu_icon="cast", default_index=0)
    selected


if selected == 'Home':
    st.write('Home')


elif selected == 'File Upload':
#^ setup file upload
    uploaded_file=st.sidebar.file_uploader(label='Upload File in .csv or .xlsx format',type=['csv','xlsx'])
    
    global df

    if uploaded_file is not None:
        st.sidebar.success('Successfully uploaded file')
        try:
            df=pd.read_csv(uploaded_file)
            
        except Exception as e:
            print(e)

            try:
                wb = load_workbook(filename=uploaded_file)
                
                sheet_names = wb.sheetnames #^ sheet_names is attribute of class Workbook and it return a list of all worksheet names
                
                df = pd.DataFrame() #! here we create a DataFrame 
                
                for sheet in sheet_names: #* here we iterate through all worksheet names and create a DataFrame for each worksheet name
                    
                    temp_df = pd.read_excel(uploaded_file, sheet_name=sheet, engine='openpyxl') #& here we create a temporary DataFrame to read the worksheet contents(SHHEET of all SHEET_NAME) and uploaded  file 
                    
                    df = df.append(temp_df) #~ here we appened the result to a DataFrame

            except Exception as e:
                print(e)
                st.sidebar.error('Please upload a valid file')
    global numeric_columns        
    try:
        st.write(df)
        numeric_columns=list(df.select_dtypes(['float','int']).columns)
    except Exception as e:
        print(e)
        st.sidebar.error('Please upload the file')

    #* Show data
    
    show_data=st.sidebar.checkbox('Show Dataset')
    if show_data:
        st.subheader('Dataset')
        st.dataframe(uploaded_file,use_container_width=True)

    #! Add a select widget to sidebar when user click on upload a file

    chart_select=st.sidebar.selectbox(
        label='Select the chart type',
        options=['Select the chart','Line Plot','Scatter Plot','Histogram Plot','Box Plot','Pie Chart']
    )

    if chart_select=='Scatter Plot':
        st.sidebar.subheader('Scatter Plot Settings')
        try:
            x_values=st.sidebar.selectbox('X axis',options=numeric_columns)
            y_values=st.sidebar.selectbox('Y axis',options=numeric_columns)
            plot=px.scatter(data_frame=df,x=x_values,y=y_values) 
            #^ display the plot
            st.plotly_charts(plot) #! not showing graph

        except Exception as e:
            print(e)





